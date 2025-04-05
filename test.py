import numpy as np
import openpyxl
import sys

from myfoos import *


#%%
class Compound:
    def __init__(self, s_formula, s_state):
        self.s_formula = s_formula
        self.s_state = s_state
        
        self.composition = decompose_formula(s_formula)
        self.M = composition_to_mass(self.composition)

        f_db = openpyxl.load_workbook("compounds_db.xlsx")
        ws = f_db["data"]
        i_index = -1
        for i in range(4, ws.max_row + 1):
            s_formula_i = ws["A" + str(i)].value
            cmp_i = decompose_formula(s_formula_i)
            
            ## сравнение составов
            isOk = True
            for element in self.composition:
                if not element in cmp_i:
                    isOk = False
                    break
                elif self.composition[element] != cmp_i[element]:
                    isOk = False
                    break
                    
            if isOk == True and self.s_state == ws["C" + str(i)].value:
                i_index = i
                break
            
        if i_index == -1:
            print("ERROR: compound not found")
            temp = input("would u like to add compound? y/n \n")
            if temp == "n":
                return
            else:
                # формула
                ws["A" + str(ws.max_row + 1)] = self.s_formula
                # название
                ws["B" + str(ws.max_row)] = input("please input name of the compound: ")
                # фаза
                ws["C" + str(ws.max_row)] = self.s_state
                # стандартная энтальпия
                ws["D" + str(ws.max_row)] = float(input("please input the dfH0_298: "))
                # стандартная энтропия
                ws["E" + str(ws.max_row)] = float(input("please input the S0_298: "))
                # коэфы
                ws["F" + str(ws.max_row)] = float(input("please input the coef Cp a: "))
                ws["G" + str(ws.max_row)] = float(input("please input the coef Cp b: "))
                ws["H" + str(ws.max_row)] = float(input("please input the coef Cp c: "))
                ws["I" + str(ws.max_row)] = float(input("please input the coef Cp c': "))
                # темпиратурный интервал
                ws["J" + str(ws.max_row)] = float(input("please input the lower limit of the Cp: "))
                ws["K" + str(ws.max_row)] = float(input("please input the higher limit of the Cp: "))
                f_db.save("compounds_db.xlsx")
                Compound(self.s_formula, self.s_state)

        self.name = ws["B" + str(i_index)].value
        self.dfH0_298 = ws["D" + str(i)].value
        self.S0_298 = ws["E" + str(i)].value

        self.cp_coeffs = []
        self.cp_powers = []

        s_cols = "FGHI"
        for ch in s_cols:
            f_coeff = ws[ch + str(i_index)].value
            i_power = ws[ch + "3"].value
            self.cp_coeffs.append(f_coeff)
            self.cp_powers.append(i_power)

        self.cp_Tmin = ws["J" + str(i_index)].value
        self.cp_Tmax = ws["K" + str(i_index)].value

        self.cp0_298 = self.calc_Cp(298.15)

    def __str__(self):
        #full_obj = self.name + " " + self.s_formula + "(" + self.s_state + ")" + "\n" "M = " + str(self.M) + "\n" + "Cp = " + str(self.cp0_298)
        return self.name + " " + self.s_formula + "(" + self.s_state + ")" + "\n" "M = " + str(self.M) + "\n" + "Cp = " + str(self.cp0_298)

    def calc_Cp(self, T):
        if T < self.cp_Tmin or T > self.cp_Tmax:
            print("ERROR T out of range")
            return -1
        Cp = 0
        for i in range(0, len(self.cp_coeffs)):
            Cp += self.cp_coeffs[i] * T**self.cp_powers[i]
        return Cp

#%%
class Reaction:
    def __init__(self, formulas, states, coeffs):
        self.coeffs = coeffs
        self.parts = []
        self.dx = 1e-3
        self.dNu = np.sum(self.coeffs)

        for i in range(len(formulas)):
            tmp = Compound(formulas[i], states[i])
            self.parts.append(tmp)

        self.drH0_298 = self.calc_heat_effect()
        self.drS0_298 = self.calc_react_entropy()
        self.drG0_298 = self.calc_gibbs()
        self.drCp_298 = self.calc_delta_Cp()

    def __str__(self):
        # Строим ур-ние р-ции
        text = ""
        s_reaction = ""
        for i in range(len(self.parts)):
            if i + 1 < len(self.parts):
                if self.coeffs[i] < 0 and self.coeffs[i + 1] > 0:
                    s_reaction += str(self.coeffs[i]) + "*" + self.parts[i].s_formula + "(" + self.parts[i].s_state + ")" + " = "
                else:
                    s_reaction += str(self.coeffs[i]) + "*" + self.parts[i].s_formula + "(" + self.parts[i].s_state + ")" + " + "
            else:
                s_reaction += str(self.coeffs[i]) + "*" + self.parts[i].s_formula + "(" + self.parts[i].s_state + ")"

        text += s_reaction + "\n"
        text += "drH0_298 = " + str(self.drH0_298) + "\n"
        text += "drS0_298 = " + str(self.drS0_298) + "\n"
        text += "drG0_298 = " + str(self.drG0_298)
        return text

    def calc_heat_effect(self, T=298.15):
        drH0_298 = 0
        for i in range(len(self.parts)):
            drH0_298 += self.parts[i].dfH0_298 * self.coeffs[i]
        if T == 298.15:
            return drH0_298

        xt = np.arange(298.15, T, self.dx if T > 298.15 else -self.dx)
        yt = [self.calc_delta_Cp(Ti) for Ti in xt]
        delta_Cp_int = integrate_trapezoid(xt, yt)
        return drH0_298 + delta_Cp_int / 1000

    def calc_react_entropy(self, T=298.15):
        drS0_298 = 0
        for i in range(len(self.parts)):
            drS0_298 += self.parts[i].S0_298 * self.coeffs[i]
        if T == 298.15:
            return drS0_298
        xt = np.arange(298.15, T, self.dx if T > 298.15 else -self.dx)
        yt = [self.calc_delta_Cp(Ti) / Ti for Ti in xt]
        delta_Cp_int = integrate_trapezoid(xt, yt)
        return drS0_298 + delta_Cp_int

    def calc_gibbs(self, T=298.15):
        drH0 = self.calc_heat_effect(T)
        drS0 = self.calc_react_entropy(T)
        drG0 = drH0 * 1000 - T * drS0
        return drG0

    def calc_delta_Cp(self, T=298.15):
        drCp = 0
        for i in range(len(self.parts)):
            drCp += self.parts[i].calc_Cp(T) * self.coeffs[i]
        return drCp

#%%
def getReaction():
    text = "1*H2(g) + 1*O2(g) = 1/2*H2O(l)"
    text = input("please insert reaction like exp: \n 1*A(g) + 2*B(g) = 3*C(g) + 4*D(g)")
    return Reaction(textToReaction(text)[0], textToReaction(text)[1], textToReaction(text)[2])


def textToReaction(text):
    import re
    from fractions import Fraction

    #print(text)

    text = text.split()

    array = []

    for i in range(len(text)):
        array.append(text[i].split("*"))
        if ["+"] in array:
            array.remove(["+"])

    eqv_sign = array.index(["="])

    for i in range(len(array)):
        if array[i][0] != "=":
            array[i][0] = Fraction(array[i][0])
        if i < eqv_sign:
            array[i][0] = array[i][0] * -1

    #print(array)

    coeffs, parts, states = [], [], []

    for i in range(len(array)):
        if array[i][0] == "=":
            continue
        coeffs.append(array[i][0])
        parts.append(array[i][1])

    #print(parts, coeffs)

    for i in range(len(parts)):
        parts[i] = re.split(r'[()]', parts[i])
        if '' in parts[i]:
            parts[i].remove('')
        states.append(parts[i][1])
        parts[i] = parts[i][0]

    return parts, states, coeffs


def getCompound():
    formula = input("Напишите формулу: ")
    state = input("введите фазу: ")
    comp = Compound(formula, state)
    return comp

xui = getCompound()

#%%
O2_test = Compound("O2", "g")
print(O2_test.s_formula)


print(O2_test.dfH0_298)


print(O2_test.cp_powers)


print(O2_test.cp0_298)


print(O2_test)

#%%
r_test = Reaction(["C4H10", "O2", "CO2", "H2O"], ["g", "g", "g", "g"], [-1, -13/2 , 4, 5])

print(r_test.drH0_298)
print(r_test)

print(r_test.calc_gibbs())

xt = np.arange(0, 10, 1e-5)
yt = xt**2
yt_int = integrate_trapezoid(xt, yt)
print(yt_int)

print(r_test.calc_heat_effect(400))
print(r_test.calc_react_entropy(400))
print(r_test.calc_gibbs(400))

print(r_test)

r_test2 = Reaction(["SO2", "O2", "SO3"],["g","g","g"],[-1,-1/2,1])

print(r_test2)

print(r_test2.calc_gibbs(298.15))

#%%
def calculate_drG(xi, react, drG0, T, P_tot, n0_s):
    prod = 1
    n_tot = 0
    for i in range(len(react.parts)):
        if react.parts[i].s_state != "g":
            continue
        nu_i = react.coeffs[i]
        n0_i = n0_s[i]
        prod *= (n0_i + nu_i * xi)**float(nu_i)
        n_tot += n0_i + nu_i*xi
    prod *= P_tot**float(react.dNu)
    prod /= n_tot**float(react.dNu)
    return drG0 + 8.314 * T * np.log(prod)

#%%
T = 900
drG0_T = r_test2.calc_gibbs(T)
print(drG0_T)

#%%
xi = np.arange(0.01, 0.99, 0.01)
drG = [calculate_drG(xi_i, r_test2, drG0_T, T, 1, [1, 0.5, 0]) for xi_i in xi]
#%%
#Построим график
import matplotlib.pyplot as plt
#%%
plt.plot(xi, drG)
plt.xlim(0, 1)
plt.plot([0,1], [0,0], color="black", linestyle="--")
